from pathlib import Path
import argparse
from typing import TypedDict, Literal
import re
import requests
import json
import numpy
import math
import datetime
import locale
import openpyxl
from openpyxl.utils import get_column_letter
from scripts.process_html_logs_extractions_to_text import (
    process_html_logs_extractions_to_text,
)


locale.setlocale(locale.LC_ALL, "pt_BR")


class AccessLog(TypedDict):
    ip: str
    port: str
    date: str


class UserAcessLogs(TypedDict):
    service: str
    identifier: str
    logs: list[AccessLog]


InfoIP_API = TypedDict(
    "InfoIP_API",
    {
        "asname": str,
        "as": str,
        "region": str,
        "city": str,
        "mobile": bool,
        "proxy": bool,
        "hosting": bool,
        "lat": float,
        "lon": float,
        "timezone": str,
        "countryCode": str,
        "status": str,
        "query": str,
    },
)


Periodo = Literal["Diurno", "Noturno"]
BILHETAGEM_KEYWORDS = ["Message Log", "Call Log", "Call Logs"]

InfoIP_Sheet = TypedDict(
    "InfoIP_Sheet",
    {
        "Alvo": str,
        "IP": str,
        "ISO_Date": str,
        "Data": str,
        "Dia da semana": str,
        "Data fuso": str,
        "IP_dono": str,
        "IP_AS": str,
        "IP_Regiao": str,
        "IP_Cidade": str,
        "IP_movel": str,
        "IP_Proxy": str,
        "IP_Hospedagem": str,
        "Periodo": Periodo,
        "Latitude": str,
        "Longitude": str,
    },
)


IP_URL = "http://ip-api.com/batch"

regex_ipv4 = r"\b(?:(?:25[0-5]|2[0-4][0-9]|1?[0-9]{1,2})\.){3}(?:25[0-5]|2[0-4][0-9]|1?[0-9]{1,2})\b"
regex_ipv6 = (
    r"(?:^|(?<=\s))(?:(?:[A-Fa-f0-9]{1,4}:){7}[A-Fa-f0-9]{1,4}|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,7}:|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,6}:[A-Fa-f0-9]{1,4}|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,5}(?::[A-Fa-f0-9]{1,4}){1,2}|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,4}(?::[A-Fa-f0-9]{1,4}){1,3}|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,3}(?::[A-Fa-f0-9]{1,4}){1,4}|"
    r"(?:[A-Fa-f0-9]{1,4}:){1,2}(?::[A-Fa-f0-9]{1,4}){1,5}|"
    r"[A-Fa-f0-9]{1,4}:(?:(?::[A-Fa-f0-9]{1,4}){1,6})|"
    r":(?:(?::[A-Fa-f0-9]{1,4}){1,7}|:))"
    r"(?:$|(?=\s))"
)


def find_index(term: str, array: list):
    index = next(index for index, line in enumerate(array) if term in line)

    return index


def find_exact_index(term: str, array: list[str]):
    index = next(
        index for index, line in enumerate(array) if term == line.replace("\n", "")
    )

    return index


def ip_parse(line: str, next_line: str):
    access_log: AccessLog = {"ip": "", "port": "", "date": ""}

    if "IP Address" in line:
        [_, full_ip] = line.replace("\n", "").split("IP Address ")

        ipv6_match = re.search(regex_ipv6, full_ip)
        ipv4_match = re.search(regex_ipv4, full_ip.split(":")[0])

        # ipv6 with port
        if "]:" in full_ip:
            [ip, port] = full_ip.replace("[", "").split("]:")
            access_log["ip"] = ip
            access_log["port"] = port
        # ipv6 without port
        elif ipv6_match:
            access_log["ip"] = full_ip
            access_log["port"] = ""
        # all ipv4 matches
        elif ipv4_match:
            # ipv4 with port
            if ":" in full_ip:
                [ip, port] = full_ip.split(":")
                access_log["ip"] = ip
                access_log["port"] = port
            # ipv4 without port
            else:
                access_log["ip"] = full_ip
                access_log["port"] = ""

        if "Time" in next_line:
            [_, time] = next_line.replace("\n", "").split("Time ")
            access_log["date"] = time
        # print(access_log)
        return access_log
    else:
        return None


def create_userlogs(file):
    file_path = Path(file)

    user_logs: UserAcessLogs = {"service": "", "identifier": "", "logs": []}

    with open(file_path, "rt", encoding="utf-8") as fp:
        lines = fp.readlines()

        service_index = find_index("Service", lines)
        identifier_index = find_index("Account Identifier", lines)
        ips_index = find_exact_index("Ip Addresses", lines)

        [_, service] = lines[service_index].split(" ")
        [_, identifier] = (
            lines[identifier_index].replace("\n", "").split("Account Identifier ")
        )

        user_logs["service"] = service.replace("\n", "")
        user_logs["identifier"] = identifier

        for index, line in enumerate(lines):
            if index < ips_index + 1:
                continue

            if index + 1 < len(lines):
                access_log = ip_parse(line, next_line=lines[index + 1])
                if access_log:
                    user_logs["logs"].append(access_log)

        return user_logs


def get_ips_info(user_logs: UserAcessLogs):
    ips_list = [log["ip"] for log in user_logs["logs"]]

    ips_list_by_100 = numpy.array_split(ips_list, math.ceil(len(ips_list) / 100))

    ips_results: dict[str, InfoIP_API] = {}

    fields = "asname,as,region,city,mobile,proxy,hosting,lat,lon,timezone,countrycode,status,query"

    try:
        for ips in ips_list_by_100:
            body = json.dumps(ips.tolist())
            response = requests.post(IP_URL, data=body, params={"fields": fields})

            data: list[InfoIP_API] = response.json()

            for index, ip in enumerate(ips):
                ips_results[ip] = data[index]

        # with open("data.json", "w+") as fj:
        #     json.dump(ips_results, fj)

        return ips_results

    except Exception as e:
        print("error", e)

        return ips_results


def create_row_date_fields(utc_date):
    string_date = utc_date.replace("UTC", "+0000")

    date = datetime.datetime.strptime(string_date, "%Y-%m-%d %H:%M:%S %z")

    current_local_time = datetime.datetime.now().astimezone()
    current_timezone = current_local_time.tzinfo

    local_date = date.astimezone(current_timezone)

    iso_date = local_date.isoformat()
    data_hora = local_date.strftime("%d/%m/%Y %H:%M")
    dia_semana = local_date.strftime("%A")
    data_fuso = "GMT " + local_date.strftime("%z")

    hours = int(local_date.strftime("%H"))

    if hours >= 5 and hours < 22:
        periodo: Periodo = "Diurno"
    else:
        periodo: Periodo = "Noturno"

    # Diurno => entre 5h às 21:59h
    # Noturno => entre 22h às 4:59h

    return [iso_date, data_hora, dia_semana, data_fuso, periodo]


def create_logs_datalist(user_logs: UserAcessLogs, ips_results: dict[str, InfoIP_API]):

    lines: list[InfoIP_Sheet] = []

    for log in user_logs["logs"]:

        info = ips_results.get(log["ip"])

        if info is None:
            continue

        port = f":[{log.get("port")}]" if log.get("port") else ""

        [iso_date, data_hora, dia_semana, data_fuso, periodo] = create_row_date_fields(
            log["date"]
        )

        sheet_line: InfoIP_Sheet = {
            "Alvo": user_logs.get("identifier"),
            "IP": log["ip"] + port,
            "ISO_Date": iso_date,
            "Data": data_hora,
            "Dia da semana": dia_semana,
            "Data fuso": data_fuso,
            "IP_dono": info.get("asname"),
            "IP_AS": info.get("as"),
            "IP_Regiao": info.get("region"),
            "IP_Cidade": info.get("city"),
            "IP_movel": str(info.get("mobile")),
            "IP_Proxy": str(info.get("proxy")),
            "IP_Hospedagem": str(info.get("hosting")),
            "Periodo": periodo,
            "Latitude": str(info.get("lat")),
            "Longitude": str(info.get("lon")),
        }

        lines.append(sheet_line)

    if not len(lines):
        print("Não há logs de acesso para salvar na planilha")

    lines.sort(key=lambda line: line["ISO_Date"], reverse=True)

    return lines


def create_logs_sheet(lines: list[InfoIP_Sheet], path: Path, user_logs: UserAcessLogs):
    headers = []

    if not len(lines):
        return

    for key in lines[0]:
        headers.append(key)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = "Logs de acesso"

    data = [headers]

    for line in lines:
        row_fields = []

        for key in line:
            row_fields.append(line[key])

        data.append(row_fields)

    column_widths = []

    for row in data:
        ws.append(row)

    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = 1.2 * len(cell)
            else:
                column_widths += [1.2 * len(cell)]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width

    wb.save(
        path.joinpath(
            f"log-acesso-{user_logs['identifier']}-{user_logs['service']}.xlsx"
        ).resolve()
    )


def process_logfile(file: str):
    user_logs = create_userlogs(file)

    ips_results = get_ips_info(user_logs)

    # testing only
    # with open("data.json", "r+") as fj:
    #     ips_results: dict[str, InfoIP_API] = json.load(fj)

    lines = create_logs_datalist(user_logs, ips_results)

    path = Path(file).parent

    create_logs_sheet(lines, path, user_logs)


def is_file_empty(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            content = file.read()

            count = content.count("No responsive records located")

            empty_message = False
            empty_call = False

            if "Message Log\nNo responsive records located" in content:
                empty_message = True

            if "Call Logs\nNo responsive records located" in content:
                empty_call = True

            if empty_message and empty_call and count == 2:
                return True

            return False
    except:
        return False


def is_bilhetagem_file(file_path):
    if "bilhetagem" in file_path:
        return True

    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()

        return any(word in content for word in BILHETAGEM_KEYWORDS)


def create_files_list(root_path: str):
    current_path = Path(root_path)

    files: list[str] = []

    try:
        for item in current_path.rglob("*"):

            if item.name.endswith(".txt"):
                file_path = str(item.resolve())

                if is_bilhetagem_file(file_path):
                    continue

                if (
                    is_file_empty(file_path)
                    or "instructions" in item.name
                    or "preservation" in item.name
                ):
                    continue

                files.append(file_path)

        return files
    except Exception as e:
        print(f"file list error {e}")
        return files


def process_meta_text_logs(root_path: str):
    process_html_logs_extractions_to_text(root_path)

    files = create_files_list(root_path)

    print(f"Processando {len(files)} arquivos")

    for file in files:
        process_logfile(file)


def get_arguments():
    parser = argparse.ArgumentParser(description="Processador de logs da META")
    parser.add_argument(
        "--pasta_raiz",
        type=str,
        required=True,
        help="Pasta raiz contendo subpastas com arquivos html",
    )

    args = parser.parse_args()

    return args


if __name__ == "__main__":
    # python -m scripts.process_meta_text_logs --pasta_raiz
    args = get_arguments()

    root_path: str = args.pasta_raiz

    process_meta_text_logs(root_path)
