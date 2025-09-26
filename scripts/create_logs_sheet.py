from scripts.ip_api import UserAcessLogs, InfoIP_API
from typing import TypedDict, Literal
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import locale

locale.setlocale(locale.LC_ALL, "pt_BR")

Periodo = Literal["Diurno", "Noturno"]

SheetColumnNames = Literal[
    "Alvo",
    "IP",
    "ISO_Date",
    "Data",
    "Dia da semana",
    "Data fuso",
    "IP_dono",
    "IP_AS",
    "IP_Regiao",
    "IP_Cidade",
    "IP_País",
    "IP_movel",
    "IP_Proxy",
    "IP_Hospedagem",
    "Periodo",
    "Latitude",
    "Longitude",
]


InfoIP_Sheet = TypedDict(
    "InfoIP_Sheet",
    {
        "Alvo": str,
        "IP": str,
        "ISO_Date": str,
        "Data": str,
        "Dia da semana": str,
        "Data fuso": str,
        "IP_dono": str,
        "IP_AS": str,
        "IP_Regiao": str,
        "IP_Cidade": str,
        "IP_País": str,
        "IP_movel": str,
        "IP_Proxy": str,
        "IP_Hospedagem": str,
        "Periodo": Periodo,
        "Latitude": str,
        "Longitude": str,
    },
)

Column_Config = TypedDict(
    "Column_Config",
    {"new_name": str},
)


Sheet_Config = dict[SheetColumnNames, Column_Config]


def create_row_date_fields(utc_date: datetime.datetime | None):
    # always pass utc_date
    if not utc_date:
        return ["", "", "", "", ""]

    current_local_time = datetime.datetime.now().astimezone()
    current_timezone = current_local_time.tzinfo

    local_date = utc_date.astimezone(current_timezone)

    iso_date = local_date.isoformat()
    data_hora = local_date.strftime("%d/%m/%Y %H:%M")
    dia_semana = local_date.strftime("%A")
    data_fuso = "GMT " + local_date.strftime("%z")

    hours = int(local_date.strftime("%H"))

    if hours >= 5 and hours < 22:
        periodo: Periodo = "Diurno"
    else:
        periodo: Periodo = "Noturno"

    # Diurno => entre 5h às 21:59h
    # Noturno => entre 22h às 4:59h

    return [iso_date, data_hora, dia_semana, data_fuso, periodo]


def change_line_column_names(line: InfoIP_Sheet, sheet_config: Sheet_Config | None):
    if not sheet_config:
        return line

    copy: InfoIP_Sheet = {}

    for key in line:
        updated = sheet_config.get(key)

        if updated and updated.get("new_name"):
            copy[updated["new_name"]] = line[key]
        else:
            copy[key] = line[key]

    return copy


def create_logs_datalist(
    user_logs: UserAcessLogs,
    ips_results: dict[str, InfoIP_API],
    sheet_config: Sheet_Config | None = None,
):

    lines: list[InfoIP_Sheet] = []

    for log in user_logs["logs"]:

        info = ips_results.get(log["ip"])

        if info is None:
            continue

        port = f"[{log.get("port")}]" if log.get("port") else ""

        [iso_date, data_hora, dia_semana, data_fuso, periodo] = create_row_date_fields(
            log["date"]
        )

        identifier = log.get("log_identifier", user_logs.get("identifier"))

        sheet_line: InfoIP_Sheet = {
            "Alvo": identifier,
            "IP": log["ip"],
            "Porta lógica": port,
            "ISO_Date": iso_date,
            "Data": data_hora,
            "Dia da semana": dia_semana,
            "Data fuso": data_fuso,
            "IP_dono": info.get("asname", ""),
            "IP_AS": info.get("as", ""),
            "IP_Regiao": info.get("region", ""),
            "IP_Cidade": info.get("city", ""),
            "IP_País": info.get("country", ""),
            "IP_movel": str(info.get("mobile", "")),
            "IP_Proxy": str(info.get("proxy", "")),
            "IP_Hospedagem": str(info.get("hosting", "")),
            "Periodo": periodo,
            "Latitude": str(info.get("lat", "")),
            "Longitude": str(info.get("lon", "")),
        }

        sheet_line = change_line_column_names(sheet_line, sheet_config)

        lines.append(sheet_line)

    if not len(lines):
        print("Não há logs de acesso para salvar na planilha")
    else:
        lines.sort(key=lambda line: line["ISO_Date"], reverse=True)

    return lines


def create_logs_sheet(
    *,
    path: Path,
    user_logs: UserAcessLogs,
    ips_results: dict[str, InfoIP_API],
    sheet_config: Sheet_Config | None = None,
    filename="",
):
    lines = create_logs_datalist(user_logs, ips_results, sheet_config)

    headers = []

    if not len(lines):
        return

    for key in lines[0]:
        headers.append(key)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = f"Logs_{user_logs['service']}_{user_logs['identifier']}"

    data = [headers]

    for line in lines:
        row_fields = []

        for key in line:
            row_fields.append(line[key])

        data.append(row_fields)

    column_widths = []

    for row in data:
        ws.append(row)

    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = 1.2 * len(cell)
            else:
                column_widths += [1.2 * len(cell)]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width

    full_name = (
        filename + ".xlsx"
        if filename
        else f"log-acesso-{user_logs['identifier']}-{user_logs['service']}.xlsx"
    )

    wb.save(path.joinpath(full_name).resolve())
